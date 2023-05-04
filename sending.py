import telebot
import os
from time import sleep
import datetime
import wget
import logging
import openpyxl
import sqlite3
from config import TOKEN, SKOSAREV_ID

bot = telebot.TeleBot(TOKEN)

logging.basicConfig(filename="logs.log", level=logging.INFO, format=' %(asctime)s - %(levelname)s - %(message)s',
                    encoding="utf8")

LINK = "https://docs.google.com/spreadsheets/d/1tGbeevMu_7_n_pKDFjH3cNFNigClVW3v/export?format=xlsx&id=1tGbeevMu_7_n_pKDFjH3cNFNigClVW3v"

A1 = [["" for i in range(2)] for j in range(4)]
A2 = [["" for k in range(2)] for l in range(4)]
allowedusers = []

conn = sqlite3.connect("NLB.db")
cur = conn.cursor()
cur.execute("CREATE TABLE IF NOT EXISTS msgs_dlt1(message_id INT PRIMARY KEY, chat_id INT);")
cur.execute("CREATE TABLE IF NOT EXISTS msgs_dlt2(message_id INT PRIMARY KEY, chat_id INT);")
cur.execute("CREATE TABLE IF NOT EXISTS dlts(message_id INT PRIMARY KEY, chat_id INT);")
conn.commit()
conn.close()


def find_sheet(wb):
    weekday = datetime.datetime.now().weekday()
    today = datetime.datetime.today() + datetime.timedelta(days=-weekday)
    date = today.strftime("%d.%m")
    sheets = wb.sheetnames
    if date[0] == "0":
        date = date[1:]
    if date == "1.05":
        date = "2.05"
    if date == "8.05":
        date = "10.05"
    for sh in sheets:
        if date in sh:
            sheet = sh
            break
    return sheet


def delete_message():
    conn = sqlite3.connect("NLB.db")
    cur = conn.cursor()
    cur.execute("SELECT * FROM msgs_dlt1")
    messages = cur.fetchall()
    for i in messages:
        try:
            bot.delete_message(i[1], i[0])
        except:
            pass
        cur.execute("DELETE FROM msgs_dlt1 WHERE message_id=?;", (i[0],))
    conn.commit()
    conn.close()
    logging.info("msgs dlt")


def delete_message2():
    conn = sqlite3.connect("NLB.db")
    cur = conn.cursor()
    cur.execute("SELECT * FROM msgs_dlt2")
    messages = cur.fetchall()
    for i in messages:
        try:
            bot.delete_message(i[1], i[0])
        except:
            pass
        cur.execute("DELETE FROM msgs_dlt2 WHERE message_id=?;", (i[0],))
    conn.commit()
    conn.close()
    logging.info("mrng msgs dlt")


def update_schedule():
    if os.path.isfile("Schedule.xlsx"):
        os.remove("Schedule.xlsx")
    t = True
    while t:
        try:
            wget.download(LINK, "Schedule.xlsx")
            logging.info("schedule downloaded")
            t = False
        except Exception:
            logging.error("schedule download failed")
            sleep(5)
            t = True
    tmp = os.listdir()
    for item in tmp:
        if item.endswith(".tmp"):
            os.remove(item)


def update_users():
    conn = sqlite3.connect("NLB.db")
    cur = conn.cursor()
    cur.execute("SELECT * FROM users")
    usrs = cur.fetchall()
    del allowedusers[0:]
    for i in range(0, len(usrs)):
        a = usrs[i][0]
        b = usrs[i][1]
        allowedusers.append([a, b])
    logging.info("users updated")


def is_merged(r, c, sheet1):
    m = sheet1.merged_cells
    merged = False
    for i in m:
        if str(i).endswith(sheet1.cell(r, c).coordinate):
            merged = True
    return merged


def get_schedule(date):
    wb = openpyxl.load_workbook("Schedule.xlsx")
    sheet = wb[find_sheet(wb)]
    for row in range(1, sheet.max_row + 1):
        if str(sheet.cell(row, 1).value)[:10] == date:
            start_row = row
    for column in range(1, sheet.max_column + 1):
        if str(sheet.cell(2, column).value) == "11мат1":
            start_column = column
    for i in range(4):
        k = sheet.cell(start_row + i, start_column + 2).value
        if k == None or sheet.cell(start_row + i, start_column + 2).font.strike:
            k = ""
        if is_merged(start_row + i, start_column + 1, sheet):
            l = sheet.cell(start_row + i, start_column).value
            if l == None or sheet.cell(start_row + i, start_column).font.strike:
                l = ""
            if l.find("(") == -1:
                A1[i][0] = l.strip()
                A2[i][0] = l.strip()
            else:
                A1[i][0] = l[:l.find("(")].strip()
                A2[i][0] = l[:l.find("(")].strip()
            if type(k) == float:
                k = str(k)[:-2]
            A1[i][1] = k
            A2[i][1] = k
        else:
            l1 = sheet.cell(start_row + i, start_column).value
            l2 = sheet.cell(start_row + i, start_column + 1).value
            if l1 == None or sheet.cell(start_row + i, start_column).font.strike:
                l1 = ""
            if l2 == None or sheet.cell(start_row + i, start_column + 1).font.strike:
                l2 = ""
            if l1.find("(") == -1:
                A1[i][0] = l1.strip()
            else:
                A1[i][0] = l1[:l1.find("(")].strip()
            if l2.find("(") == -1:
                A2[i][0] = l2.strip()
            else:
                A2[i][0] = l2[:l2.find("(")].strip()
            if type(k) == float:
                k = str(k)[:-2]
            if k.find("/") != -1:
                A1[i][1] = k[:k.find("/")]
                A2[i][1] = k[k.find("/") + 1:]
            elif A1[i][0] != "" and A2[i][0] == "":
                A1[i][1] = k
                A2[i][1] = ""
            elif A1[i][0] == "" and A2[i][0] != "":
                A1[i][1] = ""
                A2[i][1] = k
            else:
                A1[i][1] = k
                A2[i][1] = k
        if A1[i][0] == "":
            A1[i][1] = "---"
        if A2[i][0] == "":
            A2[i][1] = "---"
    logging.info("Schedule updated")


def send_next_lesson(g1, g2):
    delete_message()
    conn = sqlite3.connect("NLB.db")
    cur = conn.cursor()
    if g1[0] != "":
        for i in allowedusers:
            if i[1] == 1:
                m1 = bot.send_message(i[0], g1[0] + " в " + g1[1]).message_id
                logging.info(f"lsn msg to {i[0]}")
                cur.execute("INSERT INTO msgs_dlt1 VALUES(?,?);", (m1, i[0]))
    if g2[0] != "":
        for i in allowedusers:
            if i[1] == 2:
                m2 = bot.send_message(i[0], g2[0] + " в " + g2[1]).message_id
                logging.info(f"lsn msg to {i[0]}")
                cur.execute("INSERT INTO msgs_dlt1 VALUES(?,?);", (m2, i[0]))
    conn.commit()
    conn.close()


def send_schedule():
    now = datetime.datetime.now()
    current_time = now.strftime("%H:%M")
    weekday = now.weekday()
    current_date = now.strftime("%Y-%m-%d")

    if (weekday == 6 or weekday <= 2) and current_time == "06:30":
        conn = sqlite3.connect("NLB.db")
        cur = conn.cursor()
        cur.execute("SELECT * FROM dlts")
        dlts = cur.fetchall()
        if len(dlts) >= 0:
            for i in dlts:
                try:
                    bot.delete_message(i[1], i[0])
                except:
                    pass
            cur.execute("DELETE FROM dlts")
            conn.commit()
            conn.close()
            logging.info("msg to skosarevv dlt")

    if (weekday >= 5 or weekday<=1) and current_time == "06:30":
        dlts = bot.send_message(SKOSAREV_ID, "!!!УРА!!!СВОБОДНЫЙ ДЕНЬ!!!").message_id
        conn = sqlite3.connect("NLB.db")
        cur = conn.cursor()
        cur.execute("INSERT INTO dlts VALUES(?,?);", (dlts, SKOSAREV_ID))
        conn.commit()
        conn.close()
        logging.info("msg for skosarevv")
        sleep(60)
    elif 4 >= weekday >= 2:
        if current_time == "06:30":
            update_schedule()
            get_schedule(current_date)
            update_users()
            conn = sqlite3.connect("NLB.db")
            cur = conn.cursor()
            for i in allowedusers:
                if i[1] == 1:
                    m1 = bot.send_message(i[0],
                                          "Расписание на сегодня:\n1. " + A1[0][0] + " [" + A1[0][
                                              1] + "]\n2. " +
                                          A1[1][0] + " [" + A1[1][1] + "]\n3. " + A1[2][0] + " [" +
                                          A1[2][1] + "]\n4. "
                                          + A1[3][0] + " [" + A1[3][1] + "]\n").message_id
                    logging.info(f"mrng msg to {i[0]}")
                    cur.execute("INSERT INTO msgs_dlt2 VALUES(?,?);", (m1, i[0]))
                elif i[1] == 2:
                    m2 = bot.send_message(i[0],
                                          "Расписание на сегодня:\n1. " + A2[0][0] + " [" + A2[0][
                                              1] + "]\n2. " +
                                          A2[1][0] + " [" + A2[1][1] + "]\n3. " + A2[2][0] + " [" +
                                          A2[2][1] + "]\n4. "
                                          + A2[3][0] + " [" + A2[3][1] + "]\n").message_id
                    logging.info(f"mrng msg to {i[0]}")
                    cur.execute("INSERT INTO msgs_dlt2 VALUES(?,?);", (m2, i[0]))
            conn.commit()
            conn.close()
            sleep(60)
        elif current_time == "08:50":
            update_schedule()
            get_schedule(current_date)
            update_users()
            g1 = A1[0]
            g2 = A2[0]
            send_next_lesson(g1, g2)
            sleep(60)
        elif current_time == "09:55":
            update_schedule()
            get_schedule(current_date)
            update_users()
            g1 = A1[1]
            g2 = A2[1]
            send_next_lesson(g1, g2)
            sleep(60)
        elif current_time == "11:15":
            update_schedule()
            get_schedule(current_date)
            update_users()
            g1 = A1[2]
            g2 = A2[2]
            send_next_lesson(g1, g2)
            sleep(60)
        elif current_time == "12:45":
            update_schedule()
            get_schedule(current_date)
            update_users()
            g1 = A1[3]
            g2 = A2[3]
            send_next_lesson(g1, g2)
            sleep(60)
        elif current_time == "14:10":
            delete_message()
            delete_message2()
            sleep(60)


while True:
    try:
        send_schedule()
        sleep(30)
    except Exception as e:
        logging.error("sending error " + str(e))
        sleep(20)
